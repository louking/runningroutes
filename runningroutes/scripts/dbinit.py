###########################################################################################
# dbinit_cli - command line database initialization
#
#       Date            Author          Reason
#       ----            ------          ------
#       12/20/18        Lou King        Create
#
#   Copyright 2018 Lou King.  All rights reserved
###########################################################################################
'''
dbinit_cli - command line database initialization
=======================================================================

'''
# standard
from os.path import join, dirname
from glob import glob
from shutil import rmtree
from csv import DictReader, DictWriter

# pypi
from openpyxl import load_workbook

# homegrown
from runningroutes import create_app
from runningroutes.models import db, Route, Interest
from runningroutes.dbinit import init_db
from runningroutes.applogging import setlogging
from runningroutes.settings import Development
from runningroutes.files import create_fidfile
from loutilities.transform import Transform

class parameterError(Exception): pass

scriptdir = dirname(__file__)
# two levels up
scriptfolder = dirname(dirname(scriptdir))
configdir = join(scriptfolder, 'config')
configfile = "runningroutes.cfg"
configpath = join(configdir, configfile)
# app.config.from_object(Production(configpath))
# appconfig = getitems(configpath, 'app')
# app.config.update(appconfig)

# create app and get configuration
app = create_app(Development(configpath), configpath)

# set up database
db.init_app(app)

# set up scoped session
with app.app_context():
    # turn on logging
    setlogging()

    # clear and initialize the database
    db.drop_all()
    db.create_all()
    init_db()
    db.session.commit()

# pick up file configuration parameters
initfolder = app.config['INIT_FOLDER']
routesfile = join(initfolder, app.config['INIT_ROUTES_FILE'])
fidfile = join(initfolder, app.config['INIT_FID_FILE'])

# create mapping from fid to filename
fidwb = load_workbook(fidfile, read_only=True)
fidsheet = fidwb['file list']
fidrows = fidsheet.iter_rows(min_row=1, values_only=True)
fid = {}

# first row is header
fidhdr = next(fidrows)
for rowvalues in fidrows:
    # for some reason this is returning 26 columns, but only 2 have data
    # header is ['fileid', 'name']
    row = dict(zip(fidhdr[0:2], rowvalues[0:2]))
    fid.update({row['fileid']:row['name']})

# read routes and import into database

## first delete all the files
interestfolders = glob(join(app.config['APP_FILE_FOLDER'], '*'))
for interestfolder in interestfolders:
    rmtree(interestfolder, ignore_errors=True)

## now read the routes
routeswb = load_workbook(routesfile, read_only=True)
routessheet = routeswb['routes']
routesrows = routessheet.iter_rows(min_row=1, values_only=True)
# set up to transform routes file to Route database record
routeshdr = 'name,distance,start location,latlng,surface,elevation gain,fileid,map,description,active'.split(',')
routedbfields = 'name,distance,start_location,latlng,surface,elevation_gain,gpx_file_id,map,description,active'.split(',')
routemapping = dict(zip(routedbfields, routeshdr))
routex = Transform(routemapping, sourceattr=False, targetattr=True)
# set up to transform path sheet to path csv file
pathxlhdr = 'lat,lng,orig ele,res,ele,cumdist(km),inserted'.split(',')
pathcsvhdr = 'lat,lng,orig_ele,res,ele,cumdist_km,inserted'.split(',')
pathmapping = dict(zip(pathcsvhdr, pathxlhdr))
pathx = Transform(pathmapping, sourceattr=False, targetattr=False)

fsrc = Interest.query.filter_by(interest='fsrc').one()

routeshdr = next(routesrows)    # pulls header
routescols = routeshdr.index(None)
for in_route in routesrows:
    # seems to be more rows than those which have data, so break out at first empty row
    if in_route[0] == None: break

    thisroute = dict(zip(routeshdr[:routescols], in_route[:routescols]))
    out_route = Route()
    routex.transform(thisroute, out_route)

    # add interest
    out_route.interest = fsrc

    # get relevant data from route file
    gpxfilename = fid[out_route.gpx_file_id]
    routepath = join(initfolder, gpxfilename+'.xlsx')
    routewb = load_workbook(routepath, read_only=True)
    gpxsheet = routewb[routewb.sheetnames[0]]
    pathsheet = routewb['path']
    turnssheet = routewb['turns']

    # turns go into database field
    turnsrows = turnssheet.iter_rows(min_row=2, values_only=True)
    turns = [r[0] if len(r) >= 1 else '' for r in turnsrows]
    out_route.turns = '\n'.join(turns)

    # create gpx file
    with app.app_context():
        gpx_fid, gpxpath = create_fidfile('fsrc', fid[out_route.gpx_file_id], 'application/octet-stream', fid=out_route.gpx_file_id)
    gpxrows = gpxsheet.iter_rows(min_row=2, values_only=True)
    gpxlines = ['{}\n'.format(r[0]) for r in gpxrows]
    with open(gpxpath, 'w') as gpxfile:
        gpxfile.writelines(gpxlines)

    # create path file
    with app.app_context():
        path_fid, pathpath = create_fidfile('fsrc', fid[out_route.gpx_file_id]+'.csv', 'text/csv')
    out_route.path_file_id = path_fid
    pathrows = pathsheet.iter_rows(min_row=1, values_only=True)
    pathhdr = next(pathrows)
    with open(pathpath, 'w') as pathfile:
        out_pathcsv = DictWriter(pathfile, pathcsvhdr)
        out_pathcsv.writeheader()
        for pathrow in pathrows:
            pathcsvrow = {}
            # make sure we don't have any missing fields
            pathrowl = list(pathrow)
            while len(pathrowl) < len(pathhdr): pathrowl += [None]
            thispathrow = dict(zip(pathhdr, pathrowl))
            pathx.transform(thispathrow, pathcsvrow)
            out_pathcsv.writerow(pathcsvrow)

    # add out_route to database
    db.session.add(out_route)

db.session.commit()
