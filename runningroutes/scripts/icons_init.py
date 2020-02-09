###########################################################################################
# icons_init - command line database initialization for icons
#
#       Date            Author          Reason
#       ----            ------          ------
#       02/06/20        Lou King        Create
#
#   Copyright 2020 Lou King.  All rights reserved
###########################################################################################
'''
icons_init - command line database initialization for icons
=======================================================================
run from 2 levels up, like python -m runningroutes.scripts.icons_init

'''
# standard
from os.path import join, dirname

# pypi
from openpyxl import load_workbook

# homegrown
from runningroutes import create_app
from runningroutes.models import db, Route, Interest, Icon, IconSubtype, IconLocation, IconMap, Location, ICON_FILE_ROUTE
from runningroutes.applogging import setlogging
from runningroutes.settings import Development
from runningroutes.geo import GmapsLoc
from loutilities.transform import Transform

class parameterError(Exception): pass

scriptdir = dirname(__file__)
# two levels up
scriptfolder = dirname(dirname(scriptdir))
configdir = join(scriptfolder, 'config')
configfile = "runningroutes.cfg"
configpath = join(configdir, configfile)
# app.config.from_object(Production(configpath))
# appconfig = getitems(configpath, 'app')
# app.config.update(appconfig)

# create app and get configuration
app = create_app(Development(configpath), configpath)

# set up database access
db.init_app(app)

# set up scoped session
with app.app_context():
    # turn on logging
    setlogging()

    # alembic downgrade -1; alembic upgrade head before processing

    # clear and initialize the tables, need to do IconLocation first to avoid foreign key error
    # see https://stackoverflow.com/a/35918731/799921
    for table in [IconLocation, Icon, IconSubtype, IconMap, Location]:
        table.__table__.drop(bind=db.engine, checkfirst=True)

    for table in [Location, IconMap, IconSubtype, Icon, IconLocation]:
        table.__table__.create(bind=db.engine)

    # maybe create special "route" for icon files to reference
    route = Route.query.filter_by(name=ICON_FILE_ROUTE).one_or_none()
    if not route:
        route = Route(name=ICON_FILE_ROUTE, description='fake route to be mapped from icon related files')
        db.session.add(route)

    db.session.commit()

# pick up file configuration parameters
initfolder = app.config['INIT_LOCATIONS_FOLDER']
locsfile = join(initfolder, app.config['INIT_LOCATIONS_FILE'])

# only interest now is fsrc
fsrc = Interest.query.filter_by(interest='fsrc').one()

# read locations and import into database
locswb = load_workbook(locsfile, read_only=True, data_only=True)

## set up icons (will need to update with icon file ids)
iconssheet = locswb['icons']
iconsrows = iconssheet.iter_rows(min_row=1, values_only=True)
iconshdr = 'Icon,URL,color,Width,Height,In List,Check Attr,Show Addr'.split(',')
iconscols = len(iconshdr)
garbage = next(iconsrows) # pop header off

# set up for google maps location management
gmaps = GmapsLoc(app.config['GMAPS_ELEV_API_KEY'])

for in_icon in iconsrows:
    # seems to be more rows than those which have data, so break out at first empty row
    if in_icon[0] == None: break

    thisicon = dict(zip(iconshdr[:iconscols], in_icon[:iconscols]))
    out_icon = Icon()
    out_icon.interest = fsrc
    out_icon.icon = thisicon['Icon']
    out_icon.color = thisicon['color']
    out_icon.isAddrShown = thisicon['Show Addr'] == 'yes'
    out_icon.isShownInTable = thisicon['In List'] == 'yes'
    out_icon.isShownOnMap = True
    db.session.add(out_icon)

subtypes = ('Running Store,Retail Shop,Tap Room,Brew Pub,Restaurant,City Facility,Fitness,Coffee Shop,Activity Shop,' +
            'Point of Interest,Library,Physical Therapy,Track').split(',')
for subtype in subtypes:
    icon_st = IconSubtype(interest=fsrc, iconsubtype=subtype)
    db.session.add(icon_st)

## flush so queries later will find something
db.session.commit()

## now read the locations
locssheet = locswb['database']
locsrows = locssheet.iter_rows(min_row=1, values_only=True)
# set up to transform routes file to Route database record
locsxhdr = 'Business Name,Icon,Business Type,Comment,Street 1,City,State,Zip,Location,Business Phone,Contact Name,Contact Email,FSRC Contact,Poked?,RRCA?,FSRC?,Installed?'.split(',')
locsdbfields = 'locname,icon,iconsubtype,addl_text,location,_ro_,_ro_,_ro_,location_popup_text,phone,contact_name,email,_ro_,_ro_,_ro_,_ro_,_ro_'.split(',')
locsmapping = dict(zip(locsdbfields, locsxhdr))
del locsmapping['_ro_']

# map location information from record into something google maps will understand
# may end up as Street 1, City, State Zip or Lat, Long
def maploc(r):
    loclist = []
    for f in ['Street 1', 'City','State']:
        if r[f]:
            loclist.append( str(r[f]) )
    loc = ', '.join(loclist)
    if r['Zip']:
        loc += ' ' + str(r['Zip'])
    # since this is init function there is no location record yet
    locdata = gmaps.get_location(loc, None, app.config['GMAPS_CACHE_LIMIT'])
    locrec = Location.query.filter_by(id=locdata['id']).one()
    return locrec

locsmapping['location'] = maploc
locsmapping['icon'] = lambda r: Icon.query.filter_by(icon=r['Icon']).one() if r['Icon'] else None
locsmapping['iconsubtype'] = lambda r: IconSubtype.query.filter_by(iconsubtype=r['Business Type']).one() if r['Business Type'] else None
locsx = Transform(locsmapping, sourceattr=False, targetattr=True)

locshdr  = next(locsrows)    # pulls header
locscols = len(locsxhdr)

for in_loc in locsrows:
    # seems to be more rows than those which have data, so break out at first empty row
    if in_loc[0] == None: break

    thisloc = dict(zip(locshdr[:locscols], in_loc[:locscols]))
    out_loc = IconLocation()

    app.logger.info('processing {}'.format(thisloc))

    locsx.transform(thisloc, out_loc)

    # add interest
    out_loc.interest = fsrc

    # add out_route to database
    db.session.add(out_loc)

db.session.commit()
